import pymysql
import openpyxl

wb = openpyxl.Workbook()
ws = wb.create_sheet("学生信息", 0)
con = pymysql.connect(host="192.168.147.129", user="root", password="", database="students", port=3306)
cursor = con.cursor(cursor=pymysql.cursors.DictCursor)  # 激活游标并带出字段名
sql = "SELECT * FROM t_student_info WHERE f_id < 30;"
cursor.execute(sql)
result = cursor.fetchall()
# print(result)
ws.append(list(result[0].keys()))
ws.auto_filter.ref = ws.dimensions  # 首行添加筛选器
for key in result:
    # print(key)
    ws.append(list(key.values()))
ws.freeze_panes = ws["A2"]  # 冻结首行
wb.save("学生信息表.xlsx")
cursor.close()
con.close()
